import os

import openpyxl

baike_path = r"保险百科-表格.xlsx"
ketang_path = r"保险小课堂-表格.xlsx"
train_txt_path = r"../data/train1.txt"

def read_excel(path):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    mr = sheet.max_row
    count = 0
    with open(train_txt_path, 'a', encoding='utf-8') as f:
        # 如果前面有内容了 补一个\r\n 不然2个文件前后粘住了
        if f.tell() > 0:
            f.write('\r\n')
        for i in range(2, mr + 1):
            # 参考preprocess里面添加隔断
            question = sheet.cell(i, 2).value.strip() + '\n'
            # 答案去除换行，否则不符合训练需要的格式
            answer = sheet.cell(i, 3).value.strip().replace('\n', '') + ('\r\n' if i != mr else '')
            f.write(question)
            f.write(answer)
            count += 1
    print(f'处理{path}共{count}个对话')
    pass


if __name__ == '__main__':
    if os.path.exists(train_txt_path):
        os.remove(train_txt_path)
    read_excel(baike_path)
    read_excel(ketang_path)
    pass
